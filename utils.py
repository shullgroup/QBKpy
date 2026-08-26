# -*- coding: utf-8 -*-
# updated version of this file is maintained at
# https://github.com/shullgroup/QBKPy/blob/main/utils.py

import numpy as np
import pandas as pd
import csv
from cycler import cycler
from matplotlib.ticker import FuncFormatter, MaxNLocator
from matplotlib import rcParams
from pathlib import Path
from openpyxl import load_workbook
from scipy.interpolate import UnivariateSpline
from scipy.signal import savgol_filter

# Shared across the library
cyclers = {'broderick':cycler(color=['#0093F5', '#F08E2C', '#000000', '#424EBD', 
                               '#B04D25', '#75CA85', '#C892D6']*3, 
                        linestyle=['-']*7 + ['--']*7 + [':']*7),
           'anisha':cycler(color=['purple', '#F28C28', '#D291BC'])}

def set_default_cycler(cycler='broderick'):   
    rcParams['axes.prop_cycle'] = cyclers[cycler]
    

def spline_smooth(y, smooth_factor=1.0):
    """
    Smooth a uniformly sampled 1-D signal using a cubic smoothing spline.

    The independent variable is assumed to be the sample index
    (0, 1, 2, ...), which is appropriate for evenly spaced data.
    This function is useful for removing quantization or staircase
    artifacts caused by insufficient storage precision.

    Parameters
    ----------
    y : array-like
        Input data values.

    smooth_factor : float, optional
        Controls the amount of smoothing. Larger values produce
        smoother results.

        Rough guidelines:
            0.0   -> exact interpolation (no smoothing)
            0.1   -> light smoothing
            1.0   -> moderate smoothing (default)
            10.0  -> strong smoothing
            100.0 -> very strong smoothing

    Returns
    -------
    numpy.ndarray
        Smoothed data evaluated at the original sample locations.

    Notes
    -----
    Internally, the spline smoothing parameter is computed as

        s = smooth_factor * len(y)

    so the behavior remains reasonably consistent as the dataset size
    changes.
    """
    y = np.asarray(y, dtype=float)
    idx = np.arange(len(y))

    s = smooth_factor * len(y)

    spline = UnivariateSpline(idx, y, k=3, s=s)

    return spline(idx)

def savgol_smooth(y, window_length=11, polyorder=3):
    """
    Smooth a 1-D signal using a Savitzky-Golay filter while preserving
    leading and trailing NaN regions.

    Any NaNs at the beginning or end of the input are ignored when fitting.
    The filter is applied only to the contiguous block of valid data between
    the first and last non-NaN values. Leading and trailing NaNs are
    retained in the output.

    Parameters
    ----------
    y : array-like
        Input data values.

    window_length : int, optional
        Length of the smoothing window in samples. Must be a positive
        odd integer and greater than `polyorder`.

    polyorder : int, optional
        Degree of the polynomial fit within each window.

    Returns
    -------
    numpy.ndarray
        Smoothed data with leading and trailing NaNs preserved.
    """
    y = np.asarray(y, dtype=float)

    # Output initialized to input values
    result = y.copy()

    valid = ~np.isnan(y)
    if not np.any(valid):
        return result

    start = np.argmax(valid)
    end = len(y) - np.argmax(valid[::-1])

    segment = y[start:end]
    n = len(segment)

    # Adjust window_length if segment is shorter than requested
    wl = min(window_length, n)
    if wl % 2 == 0:
        wl -= 1

    if wl <= polyorder:
        return result

    result[start:end] = savgol_filter(
        segment,
        window_length=wl,
        polyorder=polyorder
    )

    return result




def savgol_interpolate(y, window_length=11, polyorder=3, sig_figs=4):
    """
    Smooth a uniformly sampled 1-D signal using a Savitzky-Golay
    filter while preserving the uncertainty implied by a specified
    number of significant figures.

    The function assumes each input value is known only to the
    precision corresponding to `sig_figs` significant figures.
    Smoothing is automatically limited so that no smoothed value
    differs from the original value by more than one-half of the
    implied precision.

    This is particularly useful for removing staircase artifacts
    caused by storing data with insufficient numerical precision.

    Parameters
    ----------
    y : array-like
        Input data values.

    window_length : int, optional
        Initial Savitzky-Golay window length. Must be a positive odd
        integer greater than `polyorder`.

        Larger values produce stronger smoothing. If the resulting
        smoothed values exceed the allowable precision bounds, the
        window length is automatically reduced until an acceptable
        solution is found.

        Default is 11.

    polyorder : int, optional
        Degree of the polynomial fit within each window.

        Typical values:

            2  -> quadratic fit
            3  -> cubic fit (default)
            4  -> quartic fit

        Default is 3.

    sig_figs : int, optional
        Number of significant figures represented by the input data.

        Default is 4.

        Examples for sig_figs=4:

            137.8   -> precision = 0.1
            17.86   -> precision = 0.01
            1.786   -> precision = 0.001
            0.1786  -> precision = 0.0001

    Returns
    -------
    numpy.ndarray
        Smoothed data evaluated at the original sample locations.

    Notes
    -----
    The precision associated with each value is computed as

        precision = 10**(floor(log10(abs(y))) - sig_figs + 1)

    and the maximum permitted deviation is

        precision / 2

    The function searches from the requested window length toward
    smaller window lengths and returns the smoothest result that
    satisfies

        abs(y_smooth - y) <= precision / 2

    for every point. If no acceptable smoothing is found, the
    original data are returned unchanged.
    """

    y = np.asarray(y, dtype=float)

    if len(y) < polyorder + 2:
        return y.copy()

    # Precision implied by the specified number of significant figures
    precision = np.empty_like(y)

    nonzero = np.abs(y) > 0

    precision[nonzero] = (
        10.0 ** (
            np.floor(np.log10(np.abs(y[nonzero])))
            - sig_figs + 1
        )
    )

    # For zero values use the smallest precision represented by
    # the requested number of significant figures.
    precision[~nonzero] = 10.0 ** (-sig_figs + 1)

    tolerance = precision / 2

    # Ensure a valid odd window length
    max_window = min(window_length, len(y))
    if max_window % 2 == 0:
        max_window -= 1

    min_window = polyorder + 2
    if min_window % 2 == 0:
        min_window += 1

    # Try progressively weaker smoothing until precision limits
    # are satisfied everywhere.
    for wl in range(max_window, min_window - 1, -2):

        y_smooth = savgol_filter(
            y,
            window_length=wl,
            polyorder=polyorder,
            mode='interp'
        )

        if np.all(np.abs(y_smooth - y) <= tolerance):
            return y_smooth

    return y.copy()


from scipy.stats import linregress


def local_slope(time, temp, window=101):
    """
    Estimate the local temperature ramp rate (dT/dt) using moving
    linear regressions.

    Unlike numerical differentiation methods such as
    `numpy.gradient()`, which can be highly sensitive to noise and
    quantization artifacts, this function determines the slope within
    a sliding window by fitting a straight line to the local
    temperature-time data. The resulting slope estimate is therefore
    much more robust for piecewise-linear temperature programs such as
    those commonly encountered in DSC experiments.

    Each point is assigned the slope of a least-squares linear fit to
    the neighboring data points within the specified window. For
    points near the beginning or end of the dataset, the window is
    automatically truncated to remain within the available data range.

    Parameters
    ----------
    time : array-like
        Time values. These should generally be monotonic and have the
        same length as `temp`.

    temp : array-like
        Temperature values corresponding to `time`.

    window : int, optional
        Number of points used in each local regression.

        Larger values produce smoother slope estimates and are more
        resistant to quantization artifacts, but reduce sensitivity to
        short-duration changes in ramp rate.

        Typical values:

            21   -> tracks rapid changes in slope
            51   -> moderate smoothing
            101  -> strong smoothing (default)
            201  -> very smooth slope estimates

        The value should generally be odd so that the regression
        window is approximately centered on each point.

    Returns
    -------
    numpy.ndarray
        Array of local slope estimates having the same length as the
        input arrays.

        The units are those of `temp` divided by those of `time`
        (e.g., °C/min or °C/s).

    Notes
    -----
    For each index i, a least-squares fit of the form

        T = a + b t

    is performed over a local neighborhood centered on i. The slope
    coefficient b is returned as the local estimate of dT/dt.

    This approach is particularly effective for DSC datasets because
    temperature programs are typically composed of long linear ramp
    segments separated by relatively few changes in heating or cooling
    rate. Under these conditions, local linear regression generally
    provides more reliable ramp-rate estimates than finite-difference
    derivatives.

    Examples
    --------
    Estimate a smooth heating rate:

    >>> dTdt = local_slope(df['time'], df['temp'], window=101)

    Use the resulting slopes to identify heating, cooling, and
    isothermal segments:

    >>> df['dTdt'] = local_slope(df['time'], df['temp'])
    >>> segments = find_monotonic_segments(df)
    """

    time = np.asarray(time, dtype=float)
    temp = np.asarray(temp, dtype=float)

    n = len(time)
    half = window // 2

    slope = np.zeros(n)

    for i in range(n):

        s = max(0, i - half)
        e = min(n, i + half + 1)

        slope[i] = linregress(
            time[s:e],
            temp[s:e]
        ).slope

    return slope


def is_numeric(cell):
    '''
    Check if a cell has numeric data. Used to start reading data files.
    
    Parameters
    ----------
    cell : str
        String which may be numeric (cell in a given row)
    
    Returns
    -------
    bool
        If cell is numeric (True) or not (False)
    '''
    try:
        float(cell)
        return True
    except (ValueError, TypeError):
        return False


def first_line(path, **kwargs):
    """
    Determine the first row of numeric data for CSV or Excel files.

    Parameters
    ----------
    path : str or Path
        File to inspect.

    Optional kwargs
    ---------------
    sep : str, default None
        Delimiter for text files. If None, auto-detect.

    target_cols : list[int], default None
        Columns that must contain numeric values.

    encoding : str, default 'utf-8'
        Encoding for text files.

    sheet_name : str, int, list, tuple, default 0
        Excel sheet(s) to inspect.

    Returns
    -------
    int
        First data row for a single sheet/file.

    dict
        {sheet_name: first_data_row} when multiple sheets are
        requested.
    """

    path = Path(path)
    ext = path.suffix.lower()

    sep = kwargs.get("sep", None)
    target_cols = kwargs.get("target_cols", None)
    encoding = kwargs.get("encoding", "utf-8")
    sheet_name = kwargs.get("sheet_name", 0)

    def find_first_numeric_row(rows):
        """
        Scan iterable of rows and return first row index
        containing numeric data.
        """

        for row_idx, cells in enumerate(rows):

            cells = list(cells)

            # Skip empty rows
            if all(c is None or str(c).strip() == "" for c in cells):
                continue

            if target_cols is not None:

                if len(cells) <= max(target_cols):
                    continue

                if all(is_numeric(cells[c]) for c in target_cols):
                    return row_idx

            else:

                if any(is_numeric(c) for c in cells):
                    return row_idx

        return 0

    # ------------------------------------------------------------
    # Excel files
    # ------------------------------------------------------------
    if ext in [".xls", ".xlsx"]:

        # --------------------------------------------------------
        # Multiple sheets requested
        # --------------------------------------------------------
        if isinstance(sheet_name, (list, tuple)):
            results = {}
        
            if ext == ".xlsx":
                wb = load_workbook(
                    path,
                    read_only=True,
                    data_only=True
                )
        
                for sh in sheet_name:
                    if sh == 0:
                        ws = wb.worksheets[0]           # first worksheet
                        key = ws.title                  # use actual sheet name
                    else:
                        ws = wb[sh]
                        key = sh
        
                    results[key] = find_first_numeric_row(
                        ws.iter_rows(values_only=True)
                    )
        
            else:  # .xls
                xls = pd.ExcelFile(path)
        
                for sh in sheet_name:
                    if sh == 0:
                        actual_name = xls.sheet_names[0]
                    else:
                        actual_name = sh
        
                    df_sheet = pd.read_excel(
                        path,
                        header=None,
                        sheet_name=actual_name
                    )
        
                    results[actual_name] = find_first_numeric_row(
                        (row.tolist() for _, row in df_sheet.iterrows())
                    )
        
            return results

        # --------------------------------------------------------
        # Single sheet requested
        # --------------------------------------------------------
        else:
            # Simple fallback for .xls: load once with pandas
            df = pd.read_excel(path, header=None)
            for i, row in df.iterrows():
                cells = row.tolist()
                if not all(isinstance(x, (int, float)) for x in cells):
                    continue

                if target_cols is not None:
                    if all(is_numeric(cells[c]) for c in target_cols):
                        return int(i)
                else:
                    ws = wb.active

                return find_first_numeric_row(
                    ws.iter_rows(values_only=True)
                )

            else:  # .xls

                df = pd.read_excel(
                    path,
                    header=None,
                    sheet_name=sheet_name
                )

                return find_first_numeric_row(
                    (row.tolist()
                     for _, row in df.iterrows())
                )

    # ------------------------------------------------------------
    # Text files
    # ------------------------------------------------------------

    try:
        with open(path, "r", encoding=encoding) as f:
            f.readline()
    except UnicodeDecodeError:
        encoding = "latin-1"

    # Auto-detect delimiter
    if sep is None:

        with open(path, "r", encoding=encoding) as f:

            sample = f.read(4096)

            try:
                sep = csv.Sniffer().sniff(sample).delimiter
            except Exception:
                sep = ","

    with open(path, "r", encoding=encoding) as f:

        reader = csv.reader(f, delimiter=sep)

        for i, row in enumerate(reader):

            cleaned = [cell.strip() for cell in row]

            if not cleaned:
                continue

            if target_cols is not None:
                if len(cleaned) > max(target_cols):
                    if all(is_numeric(cleaned[c]) for c in target_cols):
                        return i
            else:

                if any(is_numeric(cell)
                       for cell in cleaned):
                    return i

    return 0


def remove_step_lines(df):
    '''
    Remove extra lines in files from multiple steps in 
    TA Instruments experiments
    
    Parameters
    ----------
    df : pd.DataFrame
        Dataframe from experimental file
    
    Returns
    -------
    df : pd.DataFrame
        Cleaned dataframe with step lines removed.
    '''

    to_drop = []
    for rows in np.arange(0,len(df)):
        
        if df.iloc[rows, 0] == '[step]':
            to_drop.extend([rows,rows+1,rows+2,rows+3])
    
    df = df.drop(to_drop).reset_index(drop=True)

    return df


def read_data_file(
    path,
    target_cols=None,
    names=None,
    skiprows=None,
    header=None,
    sep=None,
    encoding="utf-8",
    sheet_name=0,
):
    """
    Read CSV, TXT, DAT, XLS, or XLSX files and return a single DataFrame.

    Parameters
    ----------
    path : str or Path
        Input file path.

    target_cols : list[int], optional
        Columns to import.

    names : list[str], optional
        Column names to assign.

    skiprows : int or dict, optional
        If None, determined automatically with first_line().

    header : int or None, default None
        Passed to pandas.

    sep : str, optional
        Delimiter for text files.

    encoding : str, default 'utf-8'
        File encoding.

    sheet_name : str, int, list, tuple, default 0
        Excel sheet(s) to read.

    Returns
    -------
    pd.DataFrame
    """

    path = Path(path)
    ext = path.suffix.lower()

    # ------------------------------------------------------------
    # Determine where numeric data begins
    # ------------------------------------------------------------
    if skiprows is None:
        skiprows = first_line(
            path,
            sep=sep,
            target_cols=target_cols,
            sheet_name=sheet_name,
            encoding=encoding,
        )

    # ------------------------------------------------------------
    # Excel files
    # ------------------------------------------------------------
    if ext in [".xls", ".xlsx"]:

        # Multiple sheets
        if isinstance(sheet_name, (list, tuple)):

            sheets = {}

            for sh in sheet_name:

                sheets[sh] = pd.read_excel(
                    path,
                    usecols=target_cols,
                    names=names,
                    skiprows=skiprows,
                    header=header,
                    sheet_name=sh,
                )

            df = pd.concat(
                sheets.values(),
                ignore_index=True
            )

        # Single sheet
        else:

            col_map = target_cols  # [time_in, temp_in, q_in]
            
            raw = pd.read_excel(
                path,
                usecols=sorted(set(col_map)),
                skiprows=skiprows,
                header=None,
                sheet_name=sheet_name,
            )
            
            # Give columns their Excel indices as names
            raw.columns = sorted(set(col_map))
            
            df = pd.DataFrame({
                'time_in': raw[col_map[0]],
                'temp_in': raw[col_map[1]],
                'q_in':    raw[col_map[2]],
            })

    # ------------------------------------------------------------
    # Text files
    # ------------------------------------------------------------
    else:

        df = pd.read_csv(
            path,
            sep=sep,
            usecols=target_cols,
            names=names,
            skiprows=skiprows,
            header=header,
            encoding=encoding,
        )

    return df



def downsample_points_per_decade(
    x,
    y=None,
    points_per_decade=10,
    base=10.0,
    include_endpoints=True,
):
    """
    Downsample data to approximately `points_per_decade` per logarithmic
    decade in x.

    Parameters
    ----------
    x : array-like, shape (n,)
        Independent variable. Must be strictly positive (for log scaling).
    y : None | array-like, shape (n,) or (n, m), optional
        Dependent variable(s) aligned with x. If provided with shape (n, m),
        each column is treated as a separate series sharing the same x.
    points_per_decade : int, default=10
        Target number of retained points per decade (in log-base `base`).
        If a decade contains fewer than this many original points, they are
        all kept.
    base : float, default=10.0
        Logarithm base that defines a “decade”. Use 10 for log10 decades,
        or e.g. `np.e` for natural-log-based “decades”.
    include_endpoints : bool, default=True
        If True, always include the global first and last indices, and try
        to include integer decade boundaries when present.

    Returns
    -------
    x_ds : ndarray, shape (k,)
        Downsampled x (sorted ascending).
    y_ds : None | ndarray
        Downsampled y aligned to x_ds. Returns None if y is None.
        If input y was (n, m), output is (k, m).
    idx : ndarray, shape (k,)
        Indices into the original arrays for the retained samples.

    Notes
    -----
    - This keeps actual data points; it does NOT interpolate.
    - Selection is done by choosing points closest to evenly spaced targets
      in log-space within each integer decade [d, d+1), except the last
      decade which includes its upper bound.
    - Duplicates are removed globally and order is preserved.
    """
    x = np.asarray(x)
    if np.any(x <= 0):
        raise ValueError("All x values must be > 0 for logarithmic decades.")
    if x.ndim != 1:
        raise ValueError("x must be 1-D.")

    n = x.size
    if y is not None:
        y = np.asarray(y)
        if y.shape[0] != n:
            raise ValueError(
                "x and y must have the same length along axis 0."
            )

    # Work in log-base `base`
    logx = np.log(x) / np.log(base)

    # Identify integer-decade bins covering the data
    d_min = int(np.floor(np.min(logx)))
    d_max = int(np.floor(np.max(logx)))  # inclusive starting decade

    selected_idx = []

    for d in range(d_min, d_max + 1):
        # For all but the last decade, use [d, d+1); for the last, include
        # right edge
        if d < d_max:
            mask = (logx >= d) & (logx < d + 1)
            right_endpoint_included = False
        else:
            mask = (logx >= d) & (logx <= d + 1)
            right_endpoint_included = True

        idx_in_decade = np.nonzero(mask)[0]
        if idx_in_decade.size == 0:
            continue

        # If there are already fewer points than target, keep them all
        if idx_in_decade.size <= points_per_decade:
            selected_idx.extend(idx_in_decade.tolist())
            continue

        # Target evenly spaced positions in log space for this decade
        if right_endpoint_included:
            targets = np.linspace(
                d, d + 1, points_per_decade, endpoint=False
            )
        else:
            targets = np.linspace(
                d, d + 1, points_per_decade, endpoint=False
            )

        # For each target, pick the index whose logx is closest
        logx_dec = logx[idx_in_decade]
        for t in targets:
            j = np.argmin(np.abs(logx_dec - t))
            selected_idx.append(idx_in_decade[j])

        # Optionally include decade boundaries if they exist in the data
        if include_endpoints:
            # Left boundary at d
            j_left = np.where(
                np.isclose(logx_dec, d, rtol=0, atol=1e-12)
            )[0]
            if j_left.size > 0:
                selected_idx.append(idx_in_decade[j_left[0]])
            # Right boundary at d+1 (only meaningful when last decade or
            # if a data point hits exactly)
            if right_endpoint_included:
                j_right = np.where(
                    np.isclose(logx_dec, d + 1, rtol=0, atol=1e-12)
                )[0]
                if j_right.size > 0:
                    selected_idx.append(idx_in_decade[j_right[0]])

    # Always include global endpoints if requested
    if include_endpoints:
        selected_idx.extend([0, n - 1])

    # Deduplicate while preserving order
    seen = set()
    ordered_unique_idx = []
    for i in selected_idx:
        if i not in seen:
            seen.add(i)
            ordered_unique_idx.append(i)
    ordered_unique_idx = np.array(ordered_unique_idx, dtype=int)

    # Sort indices by original index to preserve input sequence
    ordered_unique_idx.sort()

    x_ds = x[ordered_unique_idx]
    if y is None:
        y_ds = None
    else:
        y_ds = y[ordered_unique_idx, ...]

    return x_ds, y_ds, ordered_unique_idx


def downsample_df_per_decade(
    df: pd.DataFrame,
    column: str,
    points_per_decade: int = 10,
    base: float = 10.0,
    include_endpoints: bool = False,
):
    """
    Downsample DataFrame rows to approximately `points_per_decade`
    per logarithmic decade of the specified column.

    Parameters
    ----------
    df : pandas.DataFrame
        Input DataFrame.
    column : str
        Name of the column used for logarithmic downsampling.
        Values must be strictly positive.
    points_per_decade : int, default=10
        Target number of retained points per decade.
    base : float, default=10.0
        Logarithm base that defines a “decade”.
    include_endpoints : bool, default=False
        If True, always retain the first and last rows and try to
        include integer decade boundaries when present.

    Returns
    -------
    df_ds : pandas.DataFrame
        Downsampled DataFrame (rows preserved, no interpolation).
    idx : ndarray
        Integer indices of retained rows (relative to original df).

    Notes
    -----
    - Keeps actual data rows; does NOT interpolate.
    - Selection is made in log-space within each integer decade.
    - Row order is preserved.
    """

    if column not in df.columns:
        raise KeyError(f"Column '{column}' not found in DataFrame.")

    x = df[column].to_numpy()

    if np.any(x <= 0):
        raise ValueError(
            f"All values in column '{column}' must be > 0 "
            "for logarithmic decades."
        )

    # Work in log-base `base`
    logx = np.log(x) / np.log(base)

    d_min = int(np.floor(logx.min()))
    d_max = int(np.floor(logx.max()))

    selected_idx = []

    for d in range(d_min, d_max + 1):
        if d < d_max:
            mask = (logx >= d) & (logx < d + 1)
            right_inclusive = False
        else:
            mask = (logx >= d) & (logx <= d + 1)
            right_inclusive = True

        idx_in_decade = np.nonzero(mask)[0]
        if idx_in_decade.size == 0:
            continue

        # Keep all points if already sparse
        if idx_in_decade.size <= points_per_decade:
            selected_idx.extend(idx_in_decade.tolist())
            continue

        # Evenly spaced targets in log-space
        targets = np.linspace(
            d, d + 1, points_per_decade, endpoint=False
        )

        logx_dec = logx[idx_in_decade]

        for t in targets:
            j = np.argmin(np.abs(logx_dec - t))
            selected_idx.append(idx_in_decade[j])

        if include_endpoints:
            # Left decade boundary
            j_left = np.where(np.isclose(logx_dec, d, atol=1e-12))[0]
            if j_left.size:
                selected_idx.append(idx_in_decade[j_left[0]])

            # Right decade boundary (only for last decade)
            if right_inclusive:
                j_right = np.where(
                    np.isclose(logx_dec, d + 1, atol=1e-12)
                )[0]
                if j_right.size:
                    selected_idx.append(idx_in_decade[j_right[0]])

    # Always include global endpoints
    if include_endpoints and len(df):
        selected_idx.extend([0, len(df) - 1])

    # Deduplicate while preserving order
    seen = set()
    ordered_idx = []
    for i in selected_idx:
        if i not in seen:
            seen.add(i)
            ordered_idx.append(i)

    ordered_idx = np.array(ordered_idx, dtype=int)
    ordered_idx.sort()

    return df.iloc[ordered_idx].copy()




def baseline_correct(df, x_col, y_col, baseline, n=10):
    """
    Apply linear baseline correction using two anchor x values.
    The baseline is determined from the average of the n data points
    closest to each anchor value.
    """
    

    if len(df) < n:
        return df.copy()

    x1, x2 = [baseline[0], baseline[1]]

    # Find nearest points and keep only valid y values
    nearest1 = (df.iloc[(df[x_col] - x1).abs().argsort()]
                .dropna(subset=[y_col])
                .head(n))

    nearest2 = (df.iloc[(df[x_col] - x2).abs().argsort()]
                .dropna(subset=[y_col])
                .head(n))

    # If either endpoint does not have n valid points, return unchanged
    if len(nearest1) < n or len(nearest2) < n:
        return df.copy()

    # Use average x and y values for each anchor
    x_anchor = np.array([
        nearest1[x_col].mean(),
        nearest2[x_col].mean()
    ])

    y_anchor = np.array([
        nearest1[y_col].mean(),
        nearest2[y_col].mean()
    ])

    # Fit baseline line
    m, b = np.polyfit(x_anchor, y_anchor, 1)

    # Evaluate baseline
    baseline_vals = m * df[x_col] + b

    # Correct data
    df_out = df.copy()
    df_out[y_col] = df[y_col] - baseline_vals
    
    if len(baseline) ==3:
        idx_ref = (df_out[x_col] - baseline[2]).abs().idxmin()
        df_out[y_col] = df_out[y_col] - df_out.loc[idx_ref, y_col]

    return df_out


def add_scaled_right_axis(ax, factor, label, precision=2):
    """
    Create a right-side twin axis whose scale is linearly related
    to the left axis by a user-specified factor.

    Right axis value = Left axis value * factor
    """
    ax_right = ax.twinx()

    # Format ticks on the right axis
    ax_right.yaxis.set_major_formatter(
        FuncFormatter(lambda y, _: f"{y:.{precision}f}")
    )
    ax_right.set_ylabel(label)

    # --- Synchronize limits dynamically ---
    def sync_right_axis(ax):
        left_min, left_max = ax.get_ylim()
        ax_right.set_ylim(left_min * factor, left_max * factor)

    # Initial sync
    sync_right_axis(ax)

    # Sync whenever the left axis changes (zoom/pan)
    ax.callbacks.connect("ylim_changed", sync_right_axis)

    return ax_right


def calc_deriv(df, a, b):
       da = np.gradient(df[a])
       db = np.gradient(df[b])
       
       return np.divide(da, db,
                        out=np.full_like(da, np.nan, dtype=float),
                        where=db != 0)

